#!/usr/bin/env python3
"""
Evaluate chatbot using the provided Excel test log.

Usage:
  python scripts/evaluate_excel.py --input ../CBIT_Chatbot_Evaluation_Log.xlsx --sheet "Test Log" --out results/CBIT_Chatbot_Evaluation_Log_results.xlsx --endpoint http://localhost:5005

The script will:
- Read the specified sheet (default 'Test Log').
- For each row with a query, POST to Rasa REST webhook and collect the bot's reply(s).
- Optionally call /model/parse to get predicted intent (if endpoint supports it).
- Write the bot reply and predicted intent into the sheet and set Result to ✅/❌ by checking whether keywords from the "Expected Response / Submenu" appear in the bot reply (simple heuristic).
- Save results to the output Excel file and print a short summary (accuracy).
"""
import argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import requests
import re
from pathlib import Path


def extract_keywords(text):
    if not text:
        return []
    # split on non-word, keep tokens length>=4
    toks = re.split(r"\W+", text)
    kws = [t.lower() for t in toks if len(t) >= 4]
    return list(dict.fromkeys(kws))


def query_bot(endpoint, text, timeout=8):
    # webhook
    try:
        r = requests.post(f"{endpoint.rstrip('/')}/webhooks/rest/webhook", json={"sender": "eval", "message": text}, timeout=timeout)
        replies = r.json()
        bot_text = ' '.join([ (item.get('text') or '') for item in replies ])
    except Exception as e:
        bot_text = f'ERROR: {e}'

    # parse for intent
    intent_name = None
    intent_conf = None
    try:
        r2 = requests.post(f"{endpoint.rstrip('/')}/model/parse", json={"text": text}, timeout=timeout)
        j = r2.json()
        intent = j.get('intent') or {}
        intent_name = intent.get('name')
        intent_conf = intent.get('confidence')
    except Exception:
        pass

    return bot_text, intent_name, intent_conf


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--input', default='../CBIT_Chatbot_Evaluation_Log.xlsx')
    p.add_argument('--sheet', default='Test Log')
    p.add_argument('--out', default='results/CBIT_Chatbot_Evaluation_Log_results.xlsx')
    p.add_argument('--endpoint', default='http://localhost:5005')
    p.add_argument('--start-row', type=int, default=1, help='row to start scanning for queries')
    p.add_argument('--force', action='store_true', help='overwrite existing Actual System Response cells')
    args = p.parse_args()

    inp = Path(args.input)
    if not inp.exists():
        print('Input file not found:', inp)
        return

    wb = load_workbook(inp)
    if args.sheet not in wb.sheetnames:
        print('Sheet not found:', args.sheet)
        return
    ws = wb[args.sheet]

    # Find header row and column indexes by scanning first 6 rows
    header_row = None
    headers = None
    for r in range(1, 10):
        row = [c.value for c in ws[r]]
        if any(isinstance(v, str) and 'Query' in v for v in row if v):
            header_row = r
            headers = row
            break
    if header_row is None:
        # fallback: assume row 1 is header
        header_row = 1
        headers = [c.value for c in ws[1]]

    # map columns
    col_map = {}
    for idx, h in enumerate(headers, start=1):
        if not h:
            continue
        h_low = str(h).lower()
        if 'query' in h_low:
            col_map['query'] = idx
        if 'expected' in h_low:
            col_map['expected'] = idx
        if 'actual' in h_low:
            col_map['actual'] = idx
        if 'result' in h_low:
            col_map['result'] = idx

    if 'query' not in col_map:
        print('Could not find Query column in header row:', headers)
        return

    # Prepare results workbook (copy original)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    # We'll modify the loaded workbook and save to out

    start_data_row = header_row + 1
    total = 0
    matches = 0

    for row_idx in range(start_data_row, ws.max_row + 1):
        qcell = ws.cell(row=row_idx, column=col_map['query'])
        q = qcell.value
        if not q or (isinstance(q, str) and q.strip() == ''):
            continue
        total += 1
        expected = ''
        if 'expected' in col_map:
            expected = ws.cell(row=row_idx, column=col_map['expected']).value or ''
        # Check whether to skip if actual present
        if 'actual' in col_map and not args.force:
            existing = ws.cell(row=row_idx, column=col_map['actual']).value
            if existing and str(existing).strip() != '':
                # still collect intent via parse if possible
                try:
                    bot_text, intent_name, intent_conf = query_bot(args.endpoint, q)
                except Exception:
                    bot_text, intent_name, intent_conf = ('', None, None)
                # write predicted intent into next available column
                # find or create columns for predicted intent/confidence
                pass
        bot_text, intent_name, intent_conf = query_bot(args.endpoint, q)

        # write bot_text into actual column (or create it)
        if 'actual' in col_map:
            ws.cell(row=row_idx, column=col_map['actual']).value = bot_text
        else:
            # append at end
            col = ws.max_column + 1
            ws.cell(row=header_row, column=col).value = 'Actual System Response'
            ws.cell(row=row_idx, column=col).value = bot_text
            col_map['actual'] = col

        # write predicted intent/confidence in columns
        # find/create columns
        if 'pred_intent' not in col_map:
            col = ws.max_column + 1
            ws.cell(row=header_row, column=col).value = 'Predicted Intent'
            col_map['pred_intent'] = col
        ws.cell(row=row_idx, column=col_map['pred_intent']).value = intent_name

        if 'pred_conf' not in col_map:
            col = ws.max_column + 1
            ws.cell(row=header_row, column=col).value = 'Intent Confidence'
            col_map['pred_conf'] = col
        ws.cell(row=row_idx, column=col_map['pred_conf']).value = intent_conf

        # simple matching heuristic: check any keyword from expected appears in bot_text
        match = False
        if expected and isinstance(expected, str) and bot_text and isinstance(bot_text, str):
            kws = extract_keywords(expected)
            bot_l = bot_text.lower()
            for k in kws:
                if k in bot_l:
                    match = True
                    break

        # write result
        res_col = col_map.get('result')
        if not res_col:
            res_col = ws.max_column + 1
            ws.cell(row=header_row, column=res_col).value = 'Result'
            col_map['result'] = res_col

        ws.cell(row=row_idx, column=res_col).value = '✅' if match else '❌'
        if match:
            matches += 1

    # summary: add a new sheet with metrics
    summary = wb.create_sheet('Evaluation Summary')
    summary['A1'] = 'Total tested'
    summary['B1'] = total
    summary['A2'] = 'Matches'
    summary['B2'] = matches
    summary['A3'] = 'Accuracy'
    summary['B3'] = (matches / total) if total else 0

    wb.save(out_path)
    print('Wrote results to', out_path)
    print('Tested', total, 'accuracy', (matches / total) if total else None)


if __name__ == '__main__':
    main()
